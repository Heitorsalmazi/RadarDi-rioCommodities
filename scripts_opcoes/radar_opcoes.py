"""
radar_opcoes.py — View materializada do histórico de opções para o Radar
==========================================================================
Transforma o histórico completo em Parquet num arquivo compacto que o
navegador consegue carregar:

    dados_opcoes/*.parquet   ->  Python calcula  ->  opcoes_radar.xlsx

O NAVEGADOR NÃO PROCESSA HISTÓRICO
-----------------------------------
São ~1.900 séries por pregão. Um ano são 475 mil linhas; três anos, 1,4
milhão. Mandar isso para o SheetJS seria pedir para o Radar travar. Aqui
o Python faz a conta pesada uma vez por dia e entrega resultado pronto.

DUAS JANELAS DIFERENTES, E ISSO É PROPOSITAL
---------------------------------------------
    OPCOES_RECENTES     últimos 120 dias  -> gráficos e cadeia do dia
    ESTATISTICA_*       TODO o histórico  -> percentis e distribuições

Confundir as duas seria o erro mais fácil de cometer aqui: a estatística
ficaria míope justamente onde a série longa é o único valor que ela tem.
Os 120 dias são um recorte de ENTREGA, não de cálculo.

COMPARABILIDADE
---------------
Uma PUT de boi 5% fora do dinheiro faltando 90 dias não se compara com
uma 15% fora faltando 10 dias. Por isso todo agrupamento é

    mercado + tipo + faixa de moneyness + faixa de dias até o vencimento

e a métrica é o prêmio em PERCENTUAL DO FUTURO, que permite comparar
observações feitas com o boi a R$ 300 e a R$ 380.

A UNIDADE DE OBSERVAÇÃO É O PREGÃO
-----------------------------------
Um grupo tem muitos strikes no MESMO dia. Contar cada strike como uma
observação faria um único pregão parecer uma dúzia de dias de história —
e o percentil resultante descreveria o formato do smile daquele dia, não
a evolução do prêmio no tempo. Por isso cada grupo é agregado por DATA
antes de qualquer distribuição, e MIN_AMOSTRA_PERCENTIL conta PREGÕES.

REFERÊNCIA NÃO É NEGÓCIO
-------------------------
Hoje ~2% da cadeia negocia; o resto tem só preço de referência da B3. Os
dois universos são calculados e comparados separadamente, com variáveis
diferentes e jamais cruzados:

    REFERENCIA   preco_referencia
    NEGOCIADO    ultimo_preco, só onde houve negócio

premio_calculo não entra em estatística: ele é referência numa linha e
negócio na outra, e uma variável que muda de natureza no meio da amostra
não é uma variável.

O DIA ATUAL NÃO ENTRA NA PRÓPRIA REFERÊNCIA
--------------------------------------------
O percentil de hoje é medido contra os pregões ANTERIORES a hoje. Incluir
o próprio dia empurraria a distribuição na direção dele mesmo.

HISTÓRICO CURTO É HISTÓRICO CURTO
----------------------------------
A série começou agora. Com um pregão, TODOS os grupos saem
INSUFICIENTE e nenhum percentil é produzido — esse é o resultado
correto. O código não alarga faixa nem empresta observação de grupo
vizinho para fabricar um número.
"""

from __future__ import annotations

import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

VERSAO_GERADOR = "2.1.0"

# Versão do FORMATO entregue ao Radar — abas, colunas e semântica.
# Distinta da versão do gerador: o gerador pode mudar (correção de bug,
# otimização) sem que o HTML precise saber. O schema só muda quando o
# contrato com o consumidor muda, e aí o HTML tem que ser avisado.
SCHEMA_VERSION = "1.0"

JANELA_RADAR_DIAS = 120      # o que vai detalhado para o navegador
MIN_AMOSTRA_PERCENTIL = 20   # abaixo disso, percentil não é apresentado

ABA_RECENTES = "OPCOES_RECENTES"
ABA_FUT = "FUTUROS_RECENTES"
ABA_EST_PREMIO = "ESTATISTICA_PREMIO"
ABA_EST_LIQ = "ESTATISTICA_LIQUIDEZ"
ABA_EST_ATUAL = "ESTATISTICA_ATUAL"
ABA_CONTROLE = "CONTROLE"
ABA_SERIE = "SERIE_DIARIA"
ABA_SANIDADE = "SANIDADE"
ABA_METADADOS = "METADADOS"

# ─────────────────────────────────────────────────────────────────────────────
# Faixas — num lugar só
# ─────────────────────────────────────────────────────────────────────────────
# Mudar uma faixa aqui muda a estatística inteira de forma consistente.
# Espalhar esses números pelo código seria garantir que um dia dois
# trechos discordem sobre o que é "5% fora do dinheiro".

BUCKETS_MONEYNESS = [
    (-np.inf, -15.0, "< -15%"),
    (-15.0, -10.0, "-15% a -10%"),
    (-10.0, -7.5, "-10% a -7,5%"),
    (-7.5, -5.0, "-7,5% a -5%"),
    (-5.0, -2.5, "-5% a -2,5%"),
    (-2.5, 2.5, "ATM (-2,5% a +2,5%)"),
    (2.5, 5.0, "+2,5% a +5%"),
    (5.0, 7.5, "+5% a +7,5%"),
    (7.5, 10.0, "+7,5% a +10%"),
    (10.0, 15.0, "+10% a +15%"),
    (15.0, np.inf, "> +15%"),
]

BUCKETS_DTE = [
    (0, 15, "0-15"),
    (16, 30, "16-30"),
    (31, 45, "31-45"),
    (46, 60, "46-60"),
    (61, 90, "61-90"),
    (91, 120, "91-120"),
    (121, 180, "121-180"),
    (181, np.inf, "> 180"),
]

GRUPO = ["mercado_codigo", "tipo_opcao", "bucket_moneyness", "bucket_dte"]

COLS_RECENTES = [
    "data", "mercado", "mercado_codigo", "tipo_opcao", "ticker_opcao",
    "ticker_futuro", "vencimento", "data_vencimento",
    "tipo_vencimento_opcao", "mes_futuro_objeto",
    "unidade_preco", "fonte_preco_objeto",
    "strike", "preco_futuro", "tipo_preco_futuro", "dias_ate_vencimento",
    "distancia_strike_futuro", "moneyness_pct",
    "bucket_moneyness", "bucket_dte",
    "preco_abertura", "preco_minimo", "preco_maximo", "preco_medio",
    "ultimo_preco", "preco_referencia", "premio_calculo",
    "origem_premio_calculo", "premio_executavel", "premio_pct_futuro",
    "piso_put", "teto_call",
    "numero_negocios", "quantidade_contratos", "volume_financeiro",
    "contratos_abertos",
    "tem_negocio", "tem_bid", "tem_ask", "tem_open_interest",
    "tem_preco_referencia", "tem_cotacao_executavel",
    "melhor_compra_bid", "melhor_venda_ask", "spread_bid_ask", "spread_pct",
    "status_liquidez", "status_validacao_produto",
]


def _log(msg, nivel="INFO"):
    print(f"{datetime.now():%Y-%m-%d %H:%M:%S} | {nivel:<7} | {msg}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Enriquecimento
# ─────────────────────────────────────────────────────────────────────────────

def _rotular(valores: pd.Series, faixas: list,
             limite_superior_incluso: bool = False) -> pd.Series:
    """
    Aplica as faixas de forma vetorizada.

    np.select em vez de apply: com 1,4 milhão de linhas a diferença entre
    vetorizado e linha a linha é de segundos para minutos.

    O tratamento do limite superior é declarado por quem chama, e não
    adivinhado a partir dos números:

        moneyness  [lo, hi)  contínuo, faixas encostadas uma na outra
        DTE        [lo, hi]  inteiro, faixas "0-15", "16-30"

    Tentar deduzir isso do valor foi um erro — a versão anterior fazia
    int(hi) para descobrir se era inteiro, e estourava no np.inf da
    última faixa.
    """
    v = pd.to_numeric(valores, errors="coerce")
    if limite_superior_incluso:
        condicoes = [(v >= lo) & (v <= hi) for lo, hi, _ in faixas]
    else:
        condicoes = [(v >= lo) & (v < hi) for lo, hi, _ in faixas]
    rotulos = [r for _, _, r in faixas]
    return pd.Series(np.select(condicoes, rotulos, default=None),
                     index=valores.index, dtype="object")


def enriquecer(df: pd.DataFrame) -> pd.DataFrame:
    """
    Acrescenta buckets e prêmio percentual. Não altera coluna existente.
    """
    if df.empty:
        return df
    out = df.copy()
    out["bucket_moneyness"] = _rotular(out["moneyness_pct"], BUCKETS_MONEYNESS)
    out["bucket_dte"] = _rotular(out["dias_ate_vencimento"], BUCKETS_DTE,
                                 limite_superior_incluso=True)

    # Prêmio como % do futuro: é o que torna comparável uma observação
    # feita com o boi a R$ 300 e outra a R$ 380. Só quando os dois lados
    # são válidos e o futuro é positivo.
    pr = pd.to_numeric(out["premio_calculo"], errors="coerce")
    pf = pd.to_numeric(out["preco_futuro"], errors="coerce")
    out["premio_pct_futuro"] = (pr / pf * 100).where(pf > 0)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Estatística — a unidade de observação é o PREGÃO, não o strike
# ─────────────────────────────────────────────────────────────────────────────
# Este é o ponto mais delicado de toda a camada, e a primeira versão errou.
#
# Um grupo — digamos BGI PUT, 5% fora do dinheiro, 61 a 90 dias — costuma
# ter uma dúzia de strikes NO MESMO PREGÃO. Tratar cada strike como uma
# observação faz um único dia parecer doze dias de história. Com um pregão
# só, a base já produzia "amostra suficiente" e percentis de 5 a 100 — que
# não descreviam a evolução do prêmio no tempo, e sim o formato do smile
# naquele dia. É uma conclusão sobre outra coisa, apresentada com a cara
# da conclusão que se queria.
#
# A correção é agregar por DATA antes de qualquer distribuição:
#
#     pregão 1, grupo X, 14 strikes  ->  1 observação (mediana do dia)
#     pregão 2, grupo X,  3 strikes  ->  1 observação
#     ...
#     pregão N                       ->  1 observação
#
# É essa série temporal que gera o percentil. Assim um dia com grade densa
# não pesa mais que um dia com grade rala, e MIN_AMOSTRA_PERCENTIL volta a
# significar o que deveria: pregões distintos.
#
# Os dois universos nunca se cruzam. REFERENCIA usa preco_referencia;
# NEGOCIADO usa ultimo_preco. premio_calculo não entra em estatística
# nenhuma — ele muda de natureza conforme a linha (ora referência, ora
# negócio), e uma variável que troca de significado no meio da amostra não
# é uma variável.

UNIVERSOS = {
    "REFERENCIA": "preco_referencia",
    "NEGOCIADO": "ultimo_preco",
}


def somente_validados(hist: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra para os produtos com validação concluída.

    Produto PENDENTE continua sendo coletado, gravado no Parquet e
    exibido em OPCOES_RECENTES — some apenas das estatísticas, dos
    percentis e da SERIE_DIARIA. A ideia é não deixar o Radar apresentar
    como confiável um número cuja unidade ainda está sob verificação, sem
    por isso criar buraco no histórico.
    """
    if hist.empty or "status_validacao_produto" not in hist.columns:
        return hist
    return hist[hist["status_validacao_produto"] == "OK"]

GRUPO_DIA = ["data"] + GRUPO


def premio_pct(df: pd.DataFrame, universo: str) -> pd.Series:
    """Prêmio como % do futuro, para o universo pedido."""
    col = UNIVERSOS[universo]
    pr = pd.to_numeric(df.get(col), errors="coerce")
    pf = pd.to_numeric(df.get("preco_futuro"), errors="coerce")
    s = (pr / pf * 100).where(pf > 0)
    if universo == "NEGOCIADO":
        # Sem negócio não há preço negociado, por definição.
        s = s.where(df["tem_negocio"].fillna(False))
    return s


def camada_diaria(hist: pd.DataFrame, universo: str) -> pd.DataFrame:
    """
    Uma linha por grupo POR PREGÃO — o representante diário.

    É a peça que converte "muitos strikes" em "uma observação temporal".
    A mediana do dia é preferida à média porque a cadeia tem caudas
    grossas: um strike muito fora do dinheiro, com prêmio percentual
    desproporcional, distorceria a média do grupo.
    """
    hist = somente_validados(hist)
    if hist.empty:
        return pd.DataFrame()
    d = hist.copy()
    d["_pct"] = premio_pct(d, universo)
    d["_nom"] = pd.to_numeric(d.get(UNIVERSOS[universo]), errors="coerce")
    d = d[d["_pct"].notna()]
    if d.empty:
        return pd.DataFrame()

    g = d.groupby(GRUPO_DIA, dropna=True, observed=True)
    out = g.agg(
        n_strikes_dia=("_pct", "size"),
        premio_pct_media_dia=("_pct", "mean"),
        premio_pct_mediana_dia=("_pct", "median"),
        premio_nominal_mediana_dia=("_nom", "median"),
        n_negociadas_dia=("tem_negocio", lambda s: int(s.fillna(False).sum())),
    ).reset_index()
    out["pct_strikes_negociados_dia"] = (out["n_negociadas_dia"] /
                                         out["n_strikes_dia"] * 100)
    out.insert(0, "universo_preco", universo)
    return out


def _distribuicao(diaria: pd.DataFrame) -> pd.DataFrame:
    """
    Percentis sobre a série temporal de medianas diárias.

    n_pregoes — e não a contagem de linhas — é o que decide se há amostra.
    """
    if diaria.empty:
        return pd.DataFrame()
    g = diaria.groupby(GRUPO, dropna=True, observed=True)
    out = g.agg(
        n_pregoes=("data", "nunique"),
        n_observacoes=("n_strikes_dia", "sum"),
        n_com_negocio=("n_negociadas_dia", "sum"),
        premio_nominal_media=("premio_nominal_mediana_dia", "mean"),
        premio_nominal_mediana=("premio_nominal_mediana_dia", "median"),
        premio_pct_media=("premio_pct_mediana_dia", "mean"),
        premio_pct_mediana=("premio_pct_mediana_dia", "median"),
        premio_pct_desvio=("premio_pct_mediana_dia", "std"),
        premio_pct_min=("premio_pct_mediana_dia", "min"),
        premio_pct_max=("premio_pct_mediana_dia", "max"),
        p10=("premio_pct_mediana_dia", lambda s: s.quantile(0.10)),
        p25=("premio_pct_mediana_dia", lambda s: s.quantile(0.25)),
        p50=("premio_pct_mediana_dia", lambda s: s.quantile(0.50)),
        p75=("premio_pct_mediana_dia", lambda s: s.quantile(0.75)),
        p90=("premio_pct_mediana_dia", lambda s: s.quantile(0.90)),
    ).reset_index()

    out["pct_com_negocio"] = np.where(
        out["n_observacoes"] > 0,
        out["n_com_negocio"] / out["n_observacoes"] * 100, np.nan)
    out["status_amostra"] = np.where(
        out["n_pregoes"] >= MIN_AMOSTRA_PERCENTIL, "OK", "INSUFICIENTE")
    # Percentil sobre poucos pregões não descreve nada. Some.
    for c in ("p10", "p25", "p50", "p75", "p90", "premio_pct_desvio"):
        out.loc[out["status_amostra"] == "INSUFICIENTE", c] = np.nan
    return out


def estatistica_premio(hist: pd.DataFrame, ate_exclusivo: str = None
                       ) -> pd.DataFrame:
    """
    Distribuição do prêmio por grupo, nos dois universos.

    ate_exclusivo: quando informado, só usa pregões ANTERIORES a essa
        data. É o que a ESTATISTICA_ATUAL precisa — comparar hoje contra o
        que se sabia antes de hoje. Incluir o próprio dia na referência
        contra a qual ele é medido é autorreferência: o dia empurra a
        distribuição na direção dele mesmo, e mais ainda quando há poucos
        pregões.
    """
    if hist.empty:
        return pd.DataFrame()
    base = hist[hist["data"] < ate_exclusivo] if ate_exclusivo else hist
    if base.empty:
        return pd.DataFrame()

    partes = []
    for universo in UNIVERSOS:
        diaria = camada_diaria(base, universo)
        if diaria.empty:
            continue
        d = _distribuicao(diaria)
        d.insert(0, "universo_preco", universo)
        partes.append(d)
    if not partes:
        return pd.DataFrame()
    return pd.concat(partes, ignore_index=True).sort_values(
        ["universo_preco"] + GRUPO)


def serie_diaria(hist: pd.DataFrame) -> pd.DataFrame:
    """
    A camada diária inteira, para auditoria e para o Radar desenhar a
    evolução do prêmio de um grupo ao longo do tempo.

    É compacta: um grupo por dia, em vez de uma linha por strike.
    """
    partes = [camada_diaria(hist, u) for u in UNIVERSOS]
    partes = [p for p in partes if not p.empty]
    if not partes:
        return pd.DataFrame()
    return pd.concat(partes, ignore_index=True).sort_values(
        ["universo_preco", "data"] + GRUPO)


def estatistica_liquidez(hist: pd.DataFrame) -> pd.DataFrame:
    """
    Duas perguntas diferentes, respondidas separadamente:

      1. que fatia da cadeia costuma negociar num dia?
         -> media/mediana de pct_strikes_negociados_dia
      2. em quantos pregões esse grupo viu algum negócio?
         -> pregoes_com_algum_negocio / n_pregoes

    Misturar as duas produziria uma "frequência" que não é nem uma nem
    outra. As métricas transversais (por strike) continuam, mas nomeadas
    como tais.
    """
    hist = somente_validados(hist)
    if hist.empty:
        return pd.DataFrame()

    d = hist.copy()
    d["_qc"] = pd.to_numeric(d.get("quantidade_contratos"), errors="coerce")
    d["_oi"] = pd.to_numeric(d.get("contratos_abertos"), errors="coerce")
    d["_neg"] = d["tem_negocio"].fillna(False)

    # ── por grupo e por DIA ──────────────────────────────────────────────
    dia = d.groupby(GRUPO_DIA, dropna=True, observed=True).agg(
        n_strikes_dia=("ticker_opcao", "size"),
        n_strikes_negociados_dia=("_neg", "sum"),
        contratos_negociados_dia=("_qc", "sum"),
        open_interest_mediana_dia=("_oi", "median"),
    ).reset_index()
    dia["pct_strikes_negociados_dia"] = (dia["n_strikes_negociados_dia"] /
                                         dia["n_strikes_dia"] * 100)
    dia["houve_algum_negocio_dia"] = dia["n_strikes_negociados_dia"] > 0

    # ── sobre todo o histórico ───────────────────────────────────────────
    out = dia.groupby(GRUPO, dropna=True, observed=True).agg(
        n_pregoes=("data", "nunique"),
        pregoes_com_algum_negocio=("houve_algum_negocio_dia", "sum"),
        media_pct_strikes_negociados_dia=("pct_strikes_negociados_dia", "mean"),
        mediana_pct_strikes_negociados_dia=("pct_strikes_negociados_dia", "median"),
        media_contratos_negociados_dia=("contratos_negociados_dia", "mean"),
        mediana_contratos_negociados_dia=("contratos_negociados_dia", "median"),
        media_open_interest_mediana_dia=("open_interest_mediana_dia", "mean"),
        mediana_open_interest_mediana_dia=("open_interest_mediana_dia", "median"),
        n_observacoes_strikes=("n_strikes_dia", "sum"),
    ).reset_index()
    out["pct_pregoes_com_algum_negocio"] = (out["pregoes_com_algum_negocio"] /
                                            out["n_pregoes"] * 100)

    # Frequências transversais, por strike — úteis, mas outra pergunta
    st = d.groupby(GRUPO, dropna=True, observed=True)["status_liquidez"]
    for s in ("NEGOCIADA_NO_DIA", "COTADA", "POSICAO_ABERTA",
              "SOMENTE_REFERENCIA", "SEM_INFORMACAO"):
        freq = st.apply(lambda x, s=s: (x == s).sum() / len(x) * 100)
        out = out.merge(freq.rename(f"freq_strikes_{s.lower()}").reset_index(),
                        on=GRUPO, how="left")
    return out.sort_values(GRUPO)


# Limiares da checagem de sanidade. São de AVISO, não de bloqueio: opção
# muito dentro ou muito fora do dinheiro é legítima, e recusá-la seria
# jogar fora dado bom. O que se quer pegar é erro de UNIDADE — strike numa
# escala e futuro noutra —, que se manifesta como número absurdo.
SANIDADE_MONEYNESS_PCT = 80.0    # |moneyness| acima disso merece olhada
SANIDADE_PREMIO_PCT = 100.0      # prêmio maior que o próprio futuro
SANIDADE_RAZAO_STRIKE = 10.0     # strike 10x maior/menor que o futuro


def checar_sanidade(df: pd.DataFrame, log=_log) -> pd.DataFrame:
    """
    Procura incoerência ECONÔMICA entre strike, futuro e prêmio.

    O alvo é erro de unidade ou de escala, que aparece como número
    impossível: strike dez vezes o futuro, prêmio maior que o ativo, CALL
    valendo mais que o futuro inteiro. Uma opção muito dentro do dinheiro
    NÃO é erro, então nada é bloqueado — cada achado vira uma linha no
    relatório e um aviso no log, para revisão humana.

    Devolve as linhas suspeitas, com o motivo.
    """
    if df.empty:
        return pd.DataFrame()
    d = df.copy()
    st = pd.to_numeric(d.get("strike"), errors="coerce")
    pf = pd.to_numeric(d.get("preco_futuro"), errors="coerce")
    pr = pd.to_numeric(d.get("premio_calculo"), errors="coerce")
    mny = pd.to_numeric(d.get("moneyness_pct"), errors="coerce")

    motivos = []
    m1 = mny.abs() > SANIDADE_MONEYNESS_PCT
    m2 = (pr / pf * 100) > SANIDADE_PREMIO_PCT
    razao = (st / pf).where(pf > 0)
    m3 = (razao > SANIDADE_RAZAO_STRIKE) | (razao < 1 / SANIDADE_RAZAO_STRIKE)
    # CALL não pode valer mais que o futuro; PUT não pode valer mais que
    # o strike. São limites de arbitragem, não heurística.
    ehcall = d.get("tipo_opcao") == "CALL"
    m4 = (ehcall & (pr > pf)) | (~ehcall & (pr > st))

    for cond, texto in ((m1, f"|moneyness| > {SANIDADE_MONEYNESS_PCT:.0f}%"),
                        (m2, f"prêmio > {SANIDADE_PREMIO_PCT:.0f}% do futuro"),
                        (m3, f"strike/futuro fora de 1:{SANIDADE_RAZAO_STRIKE:.0f}"),
                        (m4, "prêmio viola limite de arbitragem")):
        motivos.append(cond.fillna(False).map({True: texto, False: ""}))
    d["_motivos"] = [" | ".join(x for x in linha if x)
                     for linha in zip(*[m.tolist() for m in motivos])]
    sus = d[d["_motivos"] != ""].copy()
    if sus.empty:
        return pd.DataFrame()

    cols = [c for c in ("data", "mercado_codigo", "tipo_opcao", "ticker_opcao",
                        "strike", "preco_futuro", "premio_calculo",
                        "moneyness_pct", "unidade_preco",
                        "status_validacao_produto") if c in sus.columns]
    out = sus[cols].copy()
    out["motivo"] = sus["_motivos"]
    por_mercado = out.groupby("mercado_codigo").size().to_dict() \
        if "mercado_codigo" in out.columns else {}
    log(f"sanidade: {len(out)} linha(s) para revisão {por_mercado}", "AVISO")
    for _, r in out.head(3).iterrows():
        log(f"   {r.get('ticker_opcao')}: {r['motivo']}", "AVISO")
    return out


def estatistica_atual(hist: pd.DataFrame) -> pd.DataFrame:
    """
    Cada opção do último pregão, medida contra o histórico ANTERIOR a ele.

    Duas leituras independentes e nunca cruzadas:

        percentil_referencia  preco_referencia de hoje contra a
                              distribuição histórica de referência
        percentil_negociado   ultimo_preco de hoje contra a distribuição
                              histórica de negócios — vazio quando não
                              houve negócio hoje, ou quando não há pregões
                              suficientes com negócio no grupo

    Comparar um negócio de hoje com uma série de preços teóricos daria um
    percentil sem significado: são variáveis diferentes.
    """
    if hist.empty:
        return pd.DataFrame()
    ultima = hist["data"].max()
    hoje = hist[hist["data"] == ultima].copy()

    base = ["mercado", "mercado_codigo", "tipo_opcao", "ticker_opcao",
            "ticker_futuro", "vencimento", "data_vencimento", "strike",
            "preco_futuro", "unidade_preco", "preco_referencia",
            "ultimo_preco", "premio_calculo", "origem_premio_calculo",
            "moneyness_pct", "bucket_moneyness", "dias_ate_vencimento",
            "bucket_dte", "status_liquidez", "tem_negocio",
            "status_validacao_produto"]
    out = hoje[[c for c in base if c in hoje.columns]].copy()
    out.insert(0, "data", ultima)
    out["premio_referencia_pct"] = premio_pct(hoje, "REFERENCIA").values
    out["premio_negociado_pct"] = premio_pct(hoje, "NEGOCIADO").values

    # Histórico estritamente anterior ao pregão de hoje
    anterior = hist[hist["data"] < ultima]

    for universo, sufixo, col_atual in (
            ("REFERENCIA", "referencia", "premio_referencia_pct"),
            ("NEGOCIADO", "negociado", "premio_negociado_pct")):
        diaria = camada_diaria(anterior, universo) if not anterior.empty \
            else pd.DataFrame()
        dist = _distribuicao(diaria) if not diaria.empty else pd.DataFrame()

        cols = {"n_pregoes": f"historico_n_pregoes_{sufixo}",
                "premio_pct_mediana": f"historico_mediana_{sufixo}",
                "p10": f"historico_p10_{sufixo}", "p25": f"historico_p25_{sufixo}",
                "p50": f"historico_p50_{sufixo}", "p75": f"historico_p75_{sufixo}",
                "p90": f"historico_p90_{sufixo}",
                "status_amostra": f"status_amostra_{sufixo}"}
        if dist.empty:
            for novo in cols.values():
                out[novo] = np.nan
            out[f"status_amostra_{sufixo}"] = "INSUFICIENTE"
            out[f"percentil_{sufixo}"] = np.nan
            continue

        out = out.merge(dist[GRUPO + list(cols)].rename(columns=cols),
                        on=GRUPO, how="left")
        out[f"status_amostra_{sufixo}"] = \
            out[f"status_amostra_{sufixo}"].fillna("INSUFICIENTE")

        # Posição na série temporal de medianas diárias, nunca nos strikes
        amostras = {chave: np.sort(g["premio_pct_mediana_dia"].to_numpy())
                    for chave, g in diaria.groupby(GRUPO, dropna=True,
                                                   observed=True)}

        def posicao(linha, amostras=amostras, sufixo=sufixo,
                    col_atual=col_atual):
            # Produto sob verificação não recebe percentil, mesmo que
            # houvesse amostra: apresentar percentil de um número cuja
            # unidade está em dúvida seria dar confiança indevida.
            if linha.get("status_validacao_produto", "OK") != "OK":
                return np.nan
            if linha[f"status_amostra_{sufixo}"] != "OK":
                return np.nan
            v = linha[col_atual]
            if pd.isna(v):
                return np.nan
            a = amostras.get(tuple(linha[c] for c in GRUPO))
            if a is None or len(a) < MIN_AMOSTRA_PERCENTIL:
                return np.nan
            return float(np.searchsorted(a, v, side="right")) / len(a) * 100

        out[f"percentil_{sufixo}"] = (out.apply(posicao, axis=1)
                                      if len(out) else np.nan)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Abas de contexto
# ─────────────────────────────────────────────────────────────────────────────

def metadados(unidades: dict, fontes: dict) -> pd.DataFrame:
    """
    O que alguém precisa saber para não interpretar a base errado.

    Vai numa aba porque o Excel viaja sozinho: quem abrir daqui a um ano,
    ou o Radar, não tem o código à mão para consultar convenções.
    """
    linhas = []
    for m in sorted(unidades):
        linhas.append({"categoria": "UNIDADE", "chave": m,
                       "valor": unidades[m],
                       "observacao": f"fonte do preço objeto: {fontes.get(m, '')}"})
    linhas += [
        {"categoria": "UNIDADE", "chave": "SJC (nota)", "valor": "USD/saca_60kg",
         "observacao": "já convertido do Mini-Sized Soybean do CME pela B3 "
                       "(fator 0,0220462). NÃO converter de novo."},
        {"categoria": "UNIDADE", "chave": "SOY (nota)", "valor": "USD/t",
         "observacao": "contrato de 34 t, referência Platts. A unidade vem da "
                       "especificação do produto; o cadastro bruto traz "
                       "TradgCcy=BRL nas opções, inconsistência da fonte."},
        {"categoria": "METODOLOGIA", "chave": "premio_pct_futuro",
         "valor": "premio_calculo / preco_futuro * 100",
         "observacao": "permite comparar observações feitas em níveis de "
                       "preço diferentes do ativo"},
        {"categoria": "SCHEMA", "chave": "schema_version",
         "valor": SCHEMA_VERSION,
         "observacao": "versão do FORMATO entregue ao Radar (abas, colunas, "
                       "semântica). Distinta de versao_gerador. O HTML deve "
                       "validar este campo antes de consumir o arquivo."},
        {"categoria": "VALIDACAO", "chave": "status_validacao_produto",
         "valor": "OK | PENDENTE",
         "observacao": "produto PENDENTE é coletado e aparece em "
                       "OPCOES_RECENTES para auditoria, mas fica FORA das "
                       "estatísticas, dos percentis e da SERIE_DIARIA. O HTML "
                       "não deve apresentar piso/teto desses como confiável."},
        {"categoria": "VALIDACAO", "chave": "aba SANIDADE",
         "valor": "linhas com incoerência econômica para revisão",
         "observacao": "avisos, não bloqueios. Procura erro de unidade ou "
                       "escala (strike 10x o futuro, prêmio acima do ativo, "
                       "violação de limite de arbitragem). Opção muito dentro "
                       "ou fora do dinheiro é legítima e não é removida."},
        {"categoria": "METODOLOGIA", "chave": "unidade de observação",
         "valor": "o PREGÃO, não o strike",
         "observacao": "cada grupo é agregado por data antes da "
                       "distribuição: muitos strikes no mesmo dia valem UMA "
                       "observação temporal (a mediana do dia). Sem isso, um "
                       "único pregão pareceria uma dúzia de dias de história."},
        {"categoria": "METODOLOGIA", "chave": "percentil_referencia",
         "valor": "posição do preco_referencia de hoje na série histórica",
         "observacao": f"exige ao menos {MIN_AMOSTRA_PERCENTIL} PREGÕES "
                       f"históricos distintos e comparáveis. Baseado "
                       f"exclusivamente em preços de referência da B3: NÃO "
                       f"representa preço executável."},
        {"categoria": "METODOLOGIA", "chave": "percentil_negociado",
         "valor": "posição do ultimo_preco de hoje na série de negócios",
         "observacao": f"exige ao menos {MIN_AMOSTRA_PERCENTIL} PREGÕES com "
                       f"negócio no grupo. Pode permanecer indisponível por "
                       f"bastante tempo: a liquidez pública das opções agro "
                       f"é baixa."},
        {"categoria": "METODOLOGIA", "chave": "exclusão do dia atual",
         "valor": "a distribuição de comparação usa data < data_atual",
         "observacao": "o pregão de hoje não entra na referência contra a "
                       "qual ele é medido — seria autorreferência."},
        {"categoria": "METODOLOGIA", "chave": "premio_calculo",
         "valor": "NÃO é usado em estatística",
         "observacao": "muda de natureza conforme a linha (referência ou "
                       "negócio). Existe para visualização. A estatística "
                       "usa preco_referencia ou ultimo_preco, separados."},
        {"categoria": "METODOLOGIA", "chave": "janela do Radar",
         "valor": f"{JANELA_RADAR_DIAS} dias corridos",
         "observacao": "vale só para OPCOES_RECENTES. As estatísticas usam "
                       "TODO o histórico do Parquet."},
        {"categoria": "PRECO", "chave": "preco_referencia",
         "valor": "preço teórico publicado pela B3",
         "observacao": "NÃO é executável. Existe para toda a cadeia, "
                       "inclusive séries sem negócio."},
        {"categoria": "PRECO", "chave": "premio_executavel",
         "valor": "midpoint de bid/ask",
         "observacao": "vazio em toda a base: os arquivos públicos da B3 não "
                       "trazem livro de ofertas. Não é estimado."},
        {"categoria": "PRECO", "chave": "universo_preco",
         "valor": "REFERENCIA | NEGOCIADO",
         "observacao": "estatísticas separadas; jamais somadas"},
        {"categoria": "LIQUIDEZ", "chave": "NEGOCIADA_NO_DIA",
         "valor": "houve negócio efetivo no pregão", "observacao": ""},
        {"categoria": "LIQUIDEZ", "chave": "COTADA",
         "valor": "há bid e/ou ask válido, mesmo sem negócio",
         "observacao": "sem fonte hoje; reservado"},
        {"categoria": "LIQUIDEZ", "chave": "POSICAO_ABERTA",
         "valor": "há contratos em aberto, sem mercado observado",
         "observacao": "posição aberta no passado NÃO significa preço "
                       "disponível hoje"},
        {"categoria": "LIQUIDEZ", "chave": "SOMENTE_REFERENCIA",
         "valor": "só o preço teórico da B3", "observacao": ""},
        {"categoria": "LIQUIDEZ", "chave": "SEM_INFORMACAO",
         "valor": "nada relevante", "observacao": ""},
    ]
    for lo, hi, rot in BUCKETS_MONEYNESS:
        linhas.append({"categoria": "BUCKET_MONEYNESS", "chave": rot,
                       "valor": f"{lo} a {hi}", "observacao": "moneyness_pct"})
    for lo, hi, rot in BUCKETS_DTE:
        linhas.append({"categoria": "BUCKET_DTE", "chave": rot,
                       "valor": f"{lo} a {hi}",
                       "observacao": "dias corridos até o vencimento"})
    return pd.DataFrame(linhas)


def controle(hist, recentes, est_premio, est_atual, segundos) -> pd.DataFrame:
    datas = sorted(hist["data"].dropna().unique()) if not hist.empty else []
    d = {
        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        # schema_version descreve o FORMATO entregue; versao_gerador, o
        # código que o produziu. O HTML valida o primeiro.
        "schema_version": SCHEMA_VERSION,
        "versao_gerador": VERSAO_GERADOR,
        "primeira_data_historico": datas[0] if datas else None,
        "ultima_data_historico": datas[-1] if datas else None,
        "total_pregoes": len(datas),
        "total_opcoes_parquet": len(hist),
        "registros_opcoes_recentes": len(recentes),
        "janela_radar_dias": JANELA_RADAR_DIAS,
        "min_amostra_percentil_PREGOES": MIN_AMOSTRA_PERCENTIL,
        "segundos_geracao": round(segundos, 1),
    }
    for m in ("BGI", "CCM", "SJC", "SOY"):
        for tp in ("PUT", "CALL"):
            sub = hist[(hist["mercado_codigo"] == m) &
                       (hist["tipo_opcao"] == tp)] if not hist.empty \
                else pd.DataFrame()
            d[f"{m.lower()}_{tp.lower()}"] = len(sub)

    d["grupos_estatisticos"] = len(est_premio)
    # Contadores POR UNIVERSO: um grupo pode ter história de referência e
    # nenhuma de negócio, e somar os dois esconderia justamente isso.
    for universo, sufixo in (("REFERENCIA", "referencia"),
                             ("NEGOCIADO", "negociado")):
        sub = (est_premio[est_premio["universo_preco"] == universo]
               if not est_premio.empty else pd.DataFrame())
        d[f"grupos_amostra_ok_{sufixo}"] = int(
            (sub["status_amostra"] == "OK").sum()) if not sub.empty else 0
        d[f"grupos_amostra_insuficiente_{sufixo}"] = int(
            (sub["status_amostra"] == "INSUFICIENTE").sum()) \
            if not sub.empty else 0
        col = f"percentil_{sufixo}"
        if not est_atual.empty and col in est_atual.columns:
            d[f"atuais_com_percentil_{sufixo}"] = int(est_atual[col].notna().sum())
            d[f"atuais_sem_percentil_{sufixo}"] = int(est_atual[col].isna().sum())
        else:
            d[f"atuais_com_percentil_{sufixo}"] = 0
            d[f"atuais_sem_percentil_{sufixo}"] = len(est_atual)
    d["max_pregoes_em_um_grupo"] = int(est_premio["n_pregoes"].max()) \
        if not est_premio.empty else 0
    if not hist.empty and "status_validacao_produto" in hist.columns:
        for m in ("BGI", "CCM", "SJC", "SOY"):
            s = hist[hist["mercado_codigo"] == m]["status_validacao_produto"]
            d[f"validacao_{m.lower()}"] = s.iloc[0] if len(s) else "-"
    return pd.DataFrame([d])


# ─────────────────────────────────────────────────────────────────────────────
# Geração
# ─────────────────────────────────────────────────────────────────────────────

def _motor_excel() -> dict:
    try:
        import xlsxwriter  # noqa: F401
        return {"engine": "xlsxwriter"}
    except ImportError:
        return {"engine": "openpyxl"}


def gerar(hist: pd.DataFrame, futuros: pd.DataFrame, destino: Path,
          unidades: dict, fontes: dict, log=_log) -> dict:
    """
    Monta o opcoes_radar.xlsx a partir do histórico completo.

    Grava num temporário, confere que abre e que as abas e contagens
    batem, e só então substitui. Se qualquer etapa falhar, o arquivo
    anterior — que estava bom — continua no lugar.
    """
    t0 = time.time()
    if hist is None or hist.empty:
        raise ValueError("histórico vazio — o Radar não é gerado a partir "
                         "de base vazia, e o arquivo anterior é preservado")

    log(f"histórico: {len(hist):,} linha(s), "
        f"{hist['data'].nunique()} pregão(ões)")
    hist = enriquecer(hist)

    # ── janela de ENTREGA (não de cálculo) ───────────────────────────────
    ultima = pd.to_datetime(hist["data"].max())
    corte = (ultima - pd.Timedelta(days=JANELA_RADAR_DIAS)).strftime("%Y-%m-%d")
    recentes = hist[hist["data"] >= corte].copy()
    recentes = recentes.reindex(
        columns=[c for c in COLS_RECENTES if c in recentes.columns])
    recentes = recentes.sort_values(
        ["data", "mercado_codigo", "tipo_opcao", "data_vencimento", "strike"])
    log(f"{ABA_RECENTES}: {len(recentes):,} linha(s) "
        f"(últimos {JANELA_RADAR_DIAS} dias, a partir de {corte})")

    fut_rec = pd.DataFrame()
    if futuros is not None and not futuros.empty:
        fut_rec = futuros[futuros["data"] >= corte].copy()
        fut_rec["unidade_preco"] = fut_rec["mercado_codigo"].map(unidades)
        fut_rec = fut_rec.sort_values(
            ["data", "mercado_codigo", "data_vencimento"])

    # ── estatística sobre TODO o histórico ───────────────────────────────
    log("calculando estatísticas sobre o histórico completo...")
    est_premio = estatistica_premio(hist)          # todo o histórico
    est_liq = estatistica_liquidez(hist)
    serie = serie_diaria(hist)                     # camada diária, auditável
    sanidade = checar_sanidade(hist, log=log)
    # A ESTATISTICA_ATUAL calcula seu próprio histórico, EXCLUINDO o
    # pregão de hoje — por isso não recebe est_premio pronto.
    est_atual = estatistica_atual(hist)
    n_pregoes = hist["data"].nunique()
    log(f"{ABA_EST_PREMIO}: {len(est_premio)} grupo(s) | "
        f"{ABA_EST_LIQ}: {len(est_liq)} | {ABA_SERIE}: {len(serie)} | "
        f"{ABA_EST_ATUAL}: {len(est_atual)}")
    if "status_validacao_produto" in hist.columns:
        pend = sorted(set(hist.loc[hist["status_validacao_produto"] != "OK",
                                   "mercado_codigo"]))
        if pend:
            log(f"produto(s) fora da estatística por validação pendente: "
                f"{', '.join(pend)} — continuam em {ABA_RECENTES} para "
                f"auditoria", "AVISO")
    if not est_premio.empty:
        ok = int((est_premio["status_amostra"] == "OK").sum())
        log(f"grupos com ao menos {MIN_AMOSTRA_PERCENTIL} pregões: {ok} de "
            f"{len(est_premio)} (a base tem {n_pregoes} pregão(ões))")
        if ok == 0:
            log(f"nenhum percentil histórico será produzido — {n_pregoes} "
                f"pregão(ões) não constitui série temporal. É o resultado "
                f"correto, não uma falha.", "AVISO")

    meta = metadados(unidades, fontes)
    ctrl = controle(hist, recentes, est_premio, est_atual, time.time() - t0)

    abas = [(ABA_RECENTES, recentes), (ABA_FUT, fut_rec),
            (ABA_EST_PREMIO, est_premio), (ABA_EST_LIQ, est_liq),
            (ABA_SERIE, serie), (ABA_EST_ATUAL, est_atual),
            (ABA_SANIDADE, sanidade), (ABA_CONTROLE, ctrl),
            (ABA_METADADOS, meta)]

    destino = Path(destino)
    temp = destino.with_name(destino.stem + ".tmp.xlsx")
    with pd.ExcelWriter(temp, **_motor_excel()) as w:
        for nome, df in abas:
            (df if not df.empty else pd.DataFrame({"aviso": [
                "sem dados suficientes nesta execução"]})
             ).to_excel(w, sheet_name=nome, index=False)

    _conferir(temp, abas)
    temp.replace(destino)

    seg = time.time() - t0
    log(f"{destino.name} gravado em {seg:.1f}s "
        f"({destino.stat().st_size/1024:.0f} KB)")
    return {"arquivo": destino, "segundos": seg,
            "recentes": len(recentes), "hist": len(hist),
            "est_premio": est_premio, "est_atual": est_atual,
            "controle": ctrl, "corte": corte}


def _conferir(temp: Path, abas: list):
    """
    Confere o arquivo gerado antes de ele virar o oficial.

    Não basta "abriu": um motor mal configurado já produziu aqui planilha
    com a primeira coluna certa e o resto vazio, sem erro nenhum. Por isso
    a checagem compara contagem de linhas aba a aba.
    """
    from openpyxl import load_workbook
    try:
        presentes = load_workbook(temp, read_only=True).sheetnames
        faltando = [n for n, _ in abas if n not in presentes]
        if faltando:
            raise RuntimeError(f"abas não gravadas: {faltando}")
        for nome, df in abas:
            if df.empty:
                continue
            lido = pd.read_excel(temp, sheet_name=nome)
            if len(lido) != len(df):
                raise RuntimeError(
                    f"aba {nome}: gravou {len(lido)} linha(s), "
                    f"esperava {len(df)}")
            if len(lido.columns) != len(df.columns):
                raise RuntimeError(
                    f"aba {nome}: {len(lido.columns)} coluna(s), "
                    f"esperava {len(df.columns)}")
    except Exception as e:
        temp.unlink(missing_ok=True)
        raise RuntimeError(f"o Radar gerado não passou na conferência: {e}")
