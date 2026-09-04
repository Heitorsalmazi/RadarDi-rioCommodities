"""
coleta_opcoes_b3.py — Banco histórico da cadeia de opções agro da B3
=====================================================================
Captura diária de TODAS as séries de opções de BGI, CCM e soja — PUT e
CALL, negociadas ou não — e acumula um histórico auditável.

ONDE O DADO MORA
----------------
    dados_opcoes/opcoes_AAAA.parquet   base completa, um arquivo por ano
    opcoes_historico.xlsx              janela recente, para você e o Radar
    temp/opcoes_b3/*.csv.gz            recorte bruto da B3, para auditoria

Por que não tudo no Excel: são ~1.900 séries por pregão, ~475 mil linhas
por ano. O limite do Excel é 1.048.576 linhas — a aba estouraria em pouco
mais de dois anos, e o histórico completo deixaria de caber justamente
quando começasse a ficar interessante para estatística.

O Parquet não tem esse teto, comprime cerca de 20x e é lido por pandas em
frações de segundo. O Excel passa a ser a JANELA de trabalho: os últimos
meses, que é o que se olha na mão. Nada é descartado — a janela é uma
vista da base, não a base.

    base completa  ->  dados_opcoes/*.parquet     (fonte da verdade)
    janela         ->  opcoes_historico.xlsx      (derivada, recriável)

Se a janela do Excel for apagada, ela se reconstrói da base. O contrário
não vale, e é por isso que o Parquet é quem manda.

NUNCA PERDER HISTÓRICO
----------------------
Coleta vazia, layout mudado, rede caída: em qualquer desses casos o
arquivo anterior fica intacto e a execução termina com erro explícito.
Não existe caminho no código que substitua a base por um dataframe vazio.

IDEMPOTÊNCIA
------------
Chave: data + mercado + tipo + ticker. Rodar duas vezes o mesmo pregão
resulta em zero registros novos.

Uso:
    python coleta_opcoes_b3.py                      # pregão mais recente
    python coleta_opcoes_b3.py --data 2026-09-02    # um pregão específico
    python coleta_opcoes_b3.py --dry-run            # não grava nada
    python coleta_opcoes_b3.py --backfill 2026-01-02 2026-09-03
    python coleta_opcoes_b3.py --refazer-excel      # só reconstrói a janela
"""

from __future__ import annotations

import argparse
import shutil
import sys
import time
import traceback
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import numpy as np
import pandas as pd

from collectors import opcoes_b3 as ob
from collectors.opcoes_b3 import FalhaColeta, LayoutMudou
import radar_opcoes as rad

BASE_DIR = Path(__file__).parent
PASTA_DADOS = BASE_DIR / "dados_opcoes"
PASTA_LOGS = BASE_DIR / "logs"
PASTA_BACKUP = BASE_DIR.parent / "backups" / "opcoes"
EXCEL = BASE_DIR.parent / "opcoes_historico.xlsx"
EXCEL_RADAR = BASE_DIR.parent / "opcoes_radar.xlsx"

ABA_HIST = "OPCOES_HISTORICO"
ABA_FUT = "FUTUROS_REFERENCIA"
ABA_RAW = "OPCOES_RAW"
ABA_CTRL = "CONTROLE"

# Quanto do histórico aparece no Excel. Medido: ~1.900 séries por pregão,
# então 3 meses ≈ 114 mil linhas, ~50s para gravar. Seis meses dobra isso e
# começa a pesar numa rotina diária — e o Radar teria de carregar o dobro
# no navegador. A base completa continua no Parquet; para olhar mais longe,
# use --janela ou leia o Parquet direto.
JANELA_MESES = 3
BACKUPS_MANTIDOS = 10

# Mínimo de séries para uma coleta ser crível. Em 02/09/2026 havia 1.872;
# se um dia vierem 12, alguma coisa quebrou na origem e não vale gravar
# por cima de um histórico bom.
MINIMO_SERIES = 100
MINIMO_MERCADOS = 2   # BGI e CCM sempre existem; soja pode faltar em teoria

_arquivo_log = None


def log(msg: str, nivel: str = "INFO"):
    linha = f"{datetime.now():%Y-%m-%d %H:%M:%S} | {nivel:<7} | {msg}"
    print(linha, flush=True)
    global _arquivo_log
    if _arquivo_log is None:
        PASTA_LOGS.mkdir(parents=True, exist_ok=True)
        _arquivo_log = PASTA_LOGS / "opcoes_b3.log"
    try:
        with open(_arquivo_log, "a", encoding="utf-8") as f:
            f.write(linha + "\n")
    except OSError:
        pass


def _log_coletor(m, n="info"):
    log(m, {"info": "INFO", "warning": "AVISO", "error": "ERRO"}.get(n, "INFO"))


# ─────────────────────────────────────────────────────────────────────────────
# Persistência — Parquet por ano, com queda para CSV comprimido
# ─────────────────────────────────────────────────────────────────────────────

def _motor_parquet() -> bool:
    try:
        import pyarrow  # noqa: F401
        return True
    except ImportError:
        try:
            import fastparquet  # noqa: F401
            return True
        except ImportError:
            return False


def _caminho_ano(ano: int, futuros: bool = False) -> Path:
    """
    Um arquivo por ano.

    Particionar por ano mantém cada arquivo em tamanho confortável e faz
    a carga diária reescrever só o ano corrente — reescrever seis anos de
    histórico todo dia seria lento e arriscaria o conjunto inteiro num
    momento de falha.
    """
    prefixo = "futuros" if futuros else "opcoes"
    ext = ".parquet" if _motor_parquet() else ".csv.gz"
    return PASTA_DADOS / f"{prefixo}_{ano}{ext}"


def _ler_ano(ano: int, futuros: bool = False) -> pd.DataFrame:
    p = _caminho_ano(ano, futuros)
    if not p.exists():
        return pd.DataFrame()
    if ".parquet" in p.suffixes:
        return pd.read_parquet(p)
    # compressão explícita: o arquivo temporário termina em .tmp e o
    # pandas não infere gzip pela extensão nesse caso
    return pd.read_csv(p, compression="gzip",
                       dtype={"data": str, "data_vencimento": str},
                       low_memory=False)


def _gravar_ano(df: pd.DataFrame, ano: int, futuros: bool = False):
    """
    Grava de forma atômica: escreve ao lado, valida lendo de volta, e só
    então substitui. Uma interrupção no meio deixa o arquivo bom no lugar.
    """
    PASTA_DADOS.mkdir(parents=True, exist_ok=True)
    destino = _caminho_ano(ano, futuros)
    temp = destino.with_suffix(destino.suffix + ".tmp")
    if ".parquet" in destino.suffixes:
        df.to_parquet(temp, index=False)
        conferido = pd.read_parquet(temp)
    else:
        df.to_csv(temp, index=False, compression="gzip")
        conferido = pd.read_csv(temp, compression="gzip", low_memory=False)
    if len(conferido) != len(df):
        temp.unlink(missing_ok=True)
        raise RuntimeError(
            f"arquivo temporário saiu com {len(conferido)} linhas, "
            f"esperava {len(df)} — o arquivo anterior foi preservado")
    temp.replace(destino)


def _carregar_base(anos=None, futuros: bool = False) -> pd.DataFrame:
    """Base completa, ou só os anos pedidos."""
    if not PASTA_DADOS.exists():
        return pd.DataFrame()
    prefixo = "futuros" if futuros else "opcoes"
    partes = []
    for p in sorted(PASTA_DADOS.glob(f"{prefixo}_*")):
        if p.name.endswith(".tmp"):
            continue
        try:
            ano = int(p.stem.split("_")[1][:4])
        except (IndexError, ValueError):
            continue
        if anos and ano not in anos:
            continue
        partes.append(_ler_ano(ano, futuros))
    return pd.concat(partes, ignore_index=True) if partes else pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# Validação — antes de qualquer gravação
# ─────────────────────────────────────────────────────────────────────────────

def validar(opcoes: pd.DataFrame, d: date) -> list:
    """
    Devolve a lista de problemas. Lista vazia = pode gravar.

    A ideia é distinguir 'dia atípico' de 'coleta quebrada'. Um pregão de
    véspera de feriado pode ter pouco negócio, e isso é dado legítimo. O
    que não pode é a cadeia inteira sumir, ou o strike vir vazio.
    """
    p = []
    if opcoes is None or opcoes.empty:
        return ["nenhuma série coletada"]

    if len(opcoes) < MINIMO_SERIES:
        p.append(f"apenas {len(opcoes)} série(s); o normal é mais de "
                 f"{MINIMO_SERIES}. Coleta suspeita.")

    mercados = set(opcoes["mercado_codigo"].dropna())
    if len(mercados) < MINIMO_MERCADOS:
        p.append(f"só {len(mercados)} mercado(s): {sorted(mercados)}")
    for esperado in ("BGI", "CCM"):
        if esperado not in mercados:
            p.append(f"{esperado} ausente — sempre deveria existir")

    tipos = set(opcoes["tipo_opcao"].dropna())
    if tipos != {"PUT", "CALL"}:
        p.append(f"tipos encontrados: {sorted(tipos)}; esperava PUT e CALL")

    for col in ("data", "mercado_codigo", "tipo_opcao", "ticker_opcao",
                "strike", "data_vencimento"):
        n = opcoes[col].isna().sum()
        if n:
            p.append(f"{n} registro(s) sem '{col}' — campo obrigatório")

    datas = set(opcoes["data"].dropna())
    if datas != {d.isoformat()}:
        p.append(f"datas divergentes no lote: {sorted(datas)[:5]}")

    dup = opcoes.duplicated(ob.CHAVE).sum()
    if dup:
        p.append(f"{dup} chave(s) repetida(s) dentro da própria coleta")

    return p


# ─────────────────────────────────────────────────────────────────────────────
# Gravação incremental
# ─────────────────────────────────────────────────────────────────────────────

def incorporar(novos: pd.DataFrame, futuros: pd.DataFrame,
               dry_run: bool = False) -> dict:
    """
    Acrescenta à base só o que ainda não existe.

    Nunca reescreve linha existente: se a chave já está lá, o registro é
    ignorado e contabilizado. Republicação da B3 com valor diferente
    aparece no log em vez de sobrescrever calado.
    """
    ano = int(str(novos["data"].iloc[0])[:4])
    atual = _ler_ano(ano)

    if atual.empty:
        acrescentar, ignorados = novos.copy(), 0
    else:
        chaves = set(map(tuple, atual[ob.CHAVE].astype(str).values))
        marca = [tuple(t) not in chaves
                 for t in novos[ob.CHAVE].astype(str).values]
        acrescentar = novos[marca].copy()
        ignorados = len(novos) - len(acrescentar)

    if not dry_run and not acrescentar.empty:
        final = (pd.concat([atual, acrescentar], ignore_index=True)
                 if not atual.empty else acrescentar)
        _gravar_ano(final, ano)

        if futuros is not None and not futuros.empty:
            fut_atual = _ler_ano(ano, futuros=True)
            chave_f = ["data", "ticker_futuro"]
            if not fut_atual.empty:
                ja = set(map(tuple, fut_atual[chave_f].astype(str).values))
                futuros = futuros[[tuple(t) not in ja for t in
                                   futuros[chave_f].astype(str).values]]
            if not futuros.empty:
                _gravar_ano(pd.concat([fut_atual, futuros], ignore_index=True)
                            if not fut_atual.empty else futuros,
                            ano, futuros=True)

    return {"novos": len(acrescentar), "ignorados": ignorados,
            "total_ano": len(atual) + len(acrescentar)}


# ─────────────────────────────────────────────────────────────────────────────
# Excel — a janela
# ─────────────────────────────────────────────────────────────────────────────

def _motor_excel() -> dict:
    """
    Qual motor usar para escrever o xlsx.

    O xlsxwriter é bem mais rápido que o openpyxl para escrever muitas
    linhas, e a janela só cresce. Sem ele instalado, openpyxl faz o
    serviço — mais devagar, mas nada deixa de funcionar.

    NÃO usar constant_memory: o pandas gera as células coluna a coluna, e
    esse modo descarta a linha assim que o cursor passa dela. O resultado
    é uma planilha com só a primeira coluna preenchida e o resto vazio —
    silenciosamente, sem erro. Já aconteceu aqui.
    """
    try:
        import xlsxwriter  # noqa: F401
        return {"engine": "xlsxwriter"}
    except ImportError:
        return {"engine": "openpyxl"}


def _backup_excel():
    """Cópia antes de mexer, com retenção limitada."""
    if not EXCEL.exists():
        return None
    PASTA_BACKUP.mkdir(parents=True, exist_ok=True)
    destino = PASTA_BACKUP / f"opcoes_historico_{date.today():%Y%m%d}.xlsx"
    if not destino.exists():
        shutil.copy2(EXCEL, destino)
    antigos = sorted(PASTA_BACKUP.glob("opcoes_historico_*.xlsx"))
    for p in antigos[:-BACKUPS_MANTIDOS]:
        p.unlink(missing_ok=True)
    return destino


def _indice_bruto() -> pd.DataFrame:
    """
    Aba OPCOES_RAW: índice dos recortes brutos guardados em disco.

    O CSV cru da B3 tem ~30 MB por pregão e centenas de milhares de linhas
    de todos os mercados — não cabe em aba de Excel e nem seria legível.
    O que fica gravado é o recorte agro comprimido, e esta aba diz onde
    cada um está, para auditoria e reprocessamento.
    """
    linhas = []
    if ob.PASTA_BRUTO.exists():
        for p in sorted(ob.PASTA_BRUTO.glob("opcoes_b3_*.csv.gz")):
            linhas.append({
                "data": p.stem.replace("opcoes_b3_", "").replace(".csv", ""),
                "arquivo": p.name,
                "caminho_relativo": f"serie_historica_updater/temp/opcoes_b3/{p.name}",
                "tamanho_kb": round(p.stat().st_size / 1024, 1),
                "gravado_em": datetime.fromtimestamp(
                    p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })
    return pd.DataFrame(linhas)


def _controle(resumo: dict) -> pd.DataFrame:
    """Aba CONTROLE — uma linha por execução, acumulando."""
    anterior = pd.DataFrame()
    if EXCEL.exists():
        try:
            anterior = pd.read_excel(EXCEL, sheet_name=ABA_CTRL)
        except Exception:
            pass
    novo = pd.DataFrame([resumo])
    return (pd.concat([anterior, novo], ignore_index=True)
            if not anterior.empty else novo)


def escrever_excel(janela_meses: int = JANELA_MESES, resumo: dict = None):
    """
    Recria o opcoes_historico.xlsx a partir da base.

    Escreve num arquivo ao lado, confere que ele abre e tem as abas
    esperadas, e só então substitui o oficial. Se qualquer etapa falhar,
    o arquivo anterior continua no lugar — e ele é descartável de todo
    modo, porque a fonte da verdade é o Parquet.
    """
    base = _carregar_base()
    if base.empty:
        log("base vazia — nada a escrever no Excel", "AVISO")
        return None

    corte = (date.today() - timedelta(days=int(janela_meses * 30.5))).isoformat()
    janela = base[base["data"] >= corte].copy()
    if janela.empty:                      # janela sem dado: mostra o que há
        janela = base.copy()
    janela = janela.sort_values(["data", "mercado_codigo", "tipo_opcao",
                                 "data_vencimento", "strike"])
    janela = janela.reindex(columns=[c for c in ob.COLUNAS
                                     if c in janela.columns])

    futuros = _carregar_base(futuros=True)
    if not futuros.empty:
        futuros = futuros[futuros["data"] >= corte].sort_values(
            ["data", "mercado_codigo", "data_vencimento"])
        futuros = futuros.reindex(
            columns=[c for c in ob.COLUNAS_FUTUROS if c in futuros.columns])

    ctrl = _controle(resumo) if resumo else (
        pd.read_excel(EXCEL, sheet_name=ABA_CTRL) if EXCEL.exists()
        else pd.DataFrame())

    temp = EXCEL.with_name(EXCEL.stem + ".tmp.xlsx")
    t0 = time.time()
    with pd.ExcelWriter(temp, **_motor_excel()) as w:
        janela.to_excel(w, sheet_name=ABA_HIST, index=False)
        if not futuros.empty:
            futuros.to_excel(w, sheet_name=ABA_FUT, index=False)
        idx = _indice_bruto()
        if not idx.empty:
            idx.to_excel(w, sheet_name=ABA_RAW, index=False)
        if ctrl is not None and not ctrl.empty:
            ctrl.to_excel(w, sheet_name=ABA_CTRL, index=False)

    # Validação: só substitui se o arquivo novo abrir mesmo
    # Conferir só "abre e tem linhas" não basta: um motor mal configurado
    # já produziu aqui uma planilha com a primeira coluna certa e todo o
    # resto vazio. Por isso a checagem olha as colunas-chave célula a
    # célula numa amostra, e compara a contagem de linhas com a origem.
    try:
        from openpyxl import load_workbook
        abas = load_workbook(temp, read_only=True).sheetnames
        if ABA_HIST not in abas:
            raise RuntimeError(f"aba '{ABA_HIST}' não foi gravada")
        conf = pd.read_excel(temp, sheet_name=ABA_HIST)
        if len(conf) != len(janela):
            raise RuntimeError(f"gravou {len(conf)} linha(s), esperava "
                               f"{len(janela)}")
        faltando = [c for c in ob.CHAVE + ["strike", "data_vencimento"]
                    if c not in conf.columns]
        if faltando:
            raise RuntimeError(f"colunas ausentes: {faltando}")
        vazias = [c for c in ob.CHAVE + ["strike"] if conf[c].isna().all()]
        if vazias:
            raise RuntimeError(f"colunas gravadas inteiramente vazias: "
                               f"{vazias}")
        if conf.duplicated(ob.CHAVE).any():
            raise RuntimeError("a planilha saiu com chave duplicada")
    except Exception as e:
        temp.unlink(missing_ok=True)
        raise RuntimeError(f"o Excel gerado não passou na conferência: {e}")

    _backup_excel()
    temp.replace(EXCEL)
    log(f"{EXCEL.name}: {len(janela):,} linha(s) na janela de "
        f"{janela_meses} meses ({janela['data'].min()} a "
        f"{janela['data'].max()}), gravado em {time.time()-t0:.1f}s")
    return EXCEL


# ─────────────────────────────────────────────────────────────────────────────
# Um pregão
# ─────────────────────────────────────────────────────────────────────────────

def processar(d: date, dry_run: bool = False) -> dict:
    """Coleta, valida e incorpora um pregão. Devolve o resumo."""
    resumo = {
        "executado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "data_b3": d.isoformat(), "status": "", "fonte": ob.API,
        "versao_coletor": ob.VERSAO_COLETOR,
        "novos": 0, "ignorados_duplicidade": 0, "erros": "", "avisos": "",
    }
    for m in ("BGI", "CCM", "SJC", "SOY"):
        resumo[f"put_{m.lower()}"] = 0
        resumo[f"call_{m.lower()}"] = 0

    try:
        r = ob.coletar(d, log=_log_coletor)
    except LayoutMudou as e:
        resumo["status"] = "ERRO_LAYOUT"
        resumo["erros"] = str(e)
        log(f"ERRO: estrutura da B3 incompatível. Histórico preservado. {e}",
            "ERRO")
        return resumo
    except FalhaColeta as e:
        resumo["status"] = "SEM_DADO"
        resumo["erros"] = str(e)
        log(f"{d:%d/%m/%Y}: {e}", "AVISO")
        return resumo

    opcoes, futuros = r["opcoes"], r["futuros"]
    problemas = validar(opcoes, d)
    if problemas:
        resumo["status"] = "REPROVADO_VALIDACAO"
        resumo["erros"] = " | ".join(problemas)
        log("VALIDAÇÃO REPROVADA — nada foi gravado, histórico preservado:",
            "ERRO")
        for p in problemas:
            log(f"   {p}", "ERRO")
        return resumo

    for m in ("BGI", "CCM", "SJC", "SOY"):
        sub = opcoes[opcoes["mercado_codigo"] == m]
        resumo[f"put_{m.lower()}"] = int((sub["tipo_opcao"] == "PUT").sum())
        resumo[f"call_{m.lower()}"] = int((sub["tipo_opcao"] == "CALL").sum())

    conta = incorporar(opcoes, futuros, dry_run=dry_run)
    resumo["novos"] = conta["novos"]
    resumo["ignorados_duplicidade"] = conta["ignorados"]
    resumo["status"] = "DRY_RUN" if dry_run else "OK"
    if r.get("status_arquivo") and "Final" not in r["status_arquivo"]:
        resumo["avisos"] = f"status do arquivo B3: {r['status_arquivo']}"
    return resumo


def ultimo_pregao(hoje: date = None) -> date:
    """
    Dia útil anterior.

    A B3 publica os arquivos do dia à noite. Uma rotina que roda de manhã
    tem que pedir D-1 — pedir o próprio dia devolveria 400 e pareceria
    falha quando é só cedo demais.
    """
    d = (hoje or date.today()) - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def reprocessar_derivados(dry_run: bool = False) -> dict:
    """
    Recalcula as colunas DERIVADAS de todo o histórico, sem rede.

    Serve para propagar uma correção de regra — como a ligação opção ->
    futuro pela raiz do ticker — para os pregões já gravados. Os dados
    brutos vêm do próprio Parquet; nada é rebaixado da B3.

    O que é recalculado: ticker_futuro, preco_futuro, tipo_preco_futuro,
    distancia_strike_futuro, moneyness_pct, dias_ate_vencimento, os sinais
    de liquidez, o status, premio_calculo e premio_executavel.

    O que NÃO é tocado: identidade, cadastro, preços, volume e posição.
    Nenhuma linha é criada ou removida — a contagem antes e depois tem que
    ser idêntica, e isso é conferido antes de gravar.
    """
    if not PASTA_DADOS.exists():
        log("nenhuma base para reprocessar", "AVISO")
        return {}

    prefixo = "opcoes"
    anos = set()
    for p in PASTA_DADOS.glob(f"{prefixo}_*"):
        if p.name.endswith(".tmp"):
            continue
        try:
            anos.add(int(p.stem.split("_")[1][:4]))
        except (IndexError, ValueError):
            pass
    if not anos:
        log("nenhum arquivo de base encontrado", "AVISO")
        return {}

    resumo = {"anos": sorted(anos), "linhas": 0, "mudou": 0}
    for ano in sorted(anos):
        antes = _ler_ano(ano)
        if antes.empty:
            continue
        futuros = _ler_ano(ano, futuros=True)
        log(f"{ano}: {len(antes):,} linha(s), "
            f"{len(futuros):,} futuro(s) de referência")

        depois = ob.recalcular_derivados(antes, futuros, log=_log_coletor)

        if len(depois) != len(antes):
            log(f"ABORTANDO {ano}: o reprocessamento mudou a contagem de "
                f"linhas ({len(antes)} -> {len(depois)}). Base preservada.",
                "ERRO")
            continue
        if depois.duplicated(ob.CHAVE).any():
            log(f"ABORTANDO {ano}: chave duplicada após reprocessar. "
                f"Base preservada.", "ERRO")
            continue

        # Conferência de que só o que devia mudar mudou
        brutas = [c for c in ("ticker_opcao", "strike", "ultimo_preco",
                              "preco_referencia", "contratos_abertos",
                              "numero_negocios", "data_vencimento")
                  if c in antes.columns and c in depois.columns]
        a = antes.sort_values(ob.CHAVE).reset_index(drop=True)
        b = depois.sort_values(ob.CHAVE).reset_index(drop=True)
        alteradas = [c for c in brutas
                     if not a[c].astype(str).equals(b[c].astype(str))]
        if alteradas:
            log(f"ABORTANDO {ano}: colunas brutas seriam alteradas "
                f"({alteradas}). Base preservada.", "ERRO")
            continue

        # Comparação com consciência de nulo: None recalculado e NaN lido do
        # Parquet são ambos "vazio", mas diferem como texto. Sem isso, a
        # segunda passada acusaria mudanças que não existem, e a rotina
        # nunca pareceria idempotente.
        def _difere_col(x, y):
            ambos_nulos = x.isna() & y.isna()
            if pd.api.types.is_numeric_dtype(x) and \
               pd.api.types.is_numeric_dtype(y):
                # Tolerância: recalcular ((strike/futuro)-1)*100 pode
                # devolver o último bit diferente do que veio do Parquet —
                # -10,0766992859032 contra -10,076699285903201. É a mesma
                # coisa a 1e-15, e tratar isso como alteração faria a
                # rotina jamais parecer idempotente.
                perto = pd.Series(
                    np.isclose(x.astype(float), y.astype(float),
                               rtol=1e-9, atol=1e-12, equal_nan=True),
                    index=x.index)
                return ~perto
            return (~ambos_nulos) & (x.astype(str) != y.astype(str))

        mudou = 0
        for c in ("ticker_futuro", "preco_futuro", "moneyness_pct",
                  "status_liquidez", "unidade_preco", "fonte_preco_objeto",
                  "tipo_vencimento_opcao", "mes_futuro_objeto"):
            if c in a.columns and c in b.columns:
                mudou += int(_difere_col(a[c], b[c]).sum())
        resumo["linhas"] += len(depois)
        resumo["mudou"] += int(mudou)

        if not dry_run:
            _gravar_ano(depois, ano)
            log(f"{ano}: gravado — {mudou:,} célula(s) derivada(s) corrigida(s)")
        else:
            log(f"{ano}: [dry-run] {mudou:,} célula(s) derivada(s) mudariam")
    return resumo


def pregoes_faltando(dias: int = 5, hoje: date = None) -> list:
    """
    Dias úteis recentes que ainda não estão na base, do mais antigo para o
    mais novo.

    A consulta é feita na base local ANTES de qualquer download. Num dia
    normal, em que só falta o último pregão, isso custa a leitura de um
    arquivo Parquet — nada de rede. Só os dias realmente ausentes são
    baixados, e cada um são três arquivos de ~30 MB.

    Sem isso, uma semana sem rodar o ATUALIZAR_HOJE viraria um buraco
    permanente na série: o coletor pediria só D-1 e os outros dias
    ficariam para trás sem ninguém perceber.
    """
    fim = ultimo_pregao(hoje)
    d, candidatos = fim, []
    while len(candidatos) < dias:
        if d.weekday() < 5:
            candidatos.append(d)
        d -= timedelta(days=1)

    anos = {c.year for c in candidatos}
    base = _carregar_base(anos=anos)
    ja = set(base["data"].astype(str)) if not base.empty else set()
    return sorted(c for c in candidatos if c.isoformat() not in ja)


def backfill(inicio: date, fim: date, dry_run: bool = False) -> list:
    """
    Preenche um intervalo, um pregão por vez.

    Data indisponível é registrada e a execução continua — feriado e fim
    de semana devolvem 400 da B3, e parar em cada um deles inviabilizaria
    qualquer intervalo longo. Roda só sob comando, nunca na rotina diária.
    """
    log(f"BACKFILL de {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}")
    resumos, d = [], inicio
    ja = set()
    base = _carregar_base(anos={a for a in range(inicio.year, fim.year + 1)})
    if not base.empty:
        ja = set(base["data"].astype(str))
        log(f"{len(ja)} pregão(ões) já na base — serão pulados")

    while d <= fim:
        if d.weekday() < 5 and d.isoformat() not in ja:
            r = processar(d, dry_run=dry_run)
            resumos.append(r)
            if r["status"] == "ERRO_LAYOUT":
                log("layout mudou — interrompendo o backfill", "ERRO")
                break
            time.sleep(1.5)     # cortesia com o servidor da B3
        d += timedelta(days=1)

    ok = [r for r in resumos if r["status"] in ("OK", "DRY_RUN")]
    log(f"BACKFILL: {len(ok)} pregão(ões) processado(s), "
        f"{sum(r['novos'] for r in ok):,} registro(s) novo(s)")
    return resumos


def gerar_radar() -> bool:
    """
    Reconstrói o opcoes_radar.xlsx a partir de TODO o histórico.

    Falha aqui não derruba a coleta: o dado do dia já está gravado no
    Parquet, que é a fonte da verdade. O Radar é uma view — sempre
    reconstruível — e é melhor ficar com a versão de ontem do que
    interromper a rotina.
    """
    try:
        hist = _carregar_base()
        if hist.empty:
            log("base vazia — o Radar não foi gerado; o arquivo anterior "
                "(se houver) permanece", "AVISO")
            return False
        rad.gerar(hist, _carregar_base(futuros=True), EXCEL_RADAR,
                  ob.UNIDADE_PRECO, ob.FONTE_PRECO_OBJETO,
                  log=lambda m, n="INFO": log(m, n))
        return True
    except Exception as e:
        log(f"Radar não foi gerado ({type(e).__name__}: {e}). O arquivo "
            f"anterior foi preservado e a base está intacta.", "ERRO")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Linha de comando
# ─────────────────────────────────────────────────────────────────────────────

def _relatorio(r: dict):
    print()
    for m, nome in (("bgi", "BGI"), ("ccm", "CCM"),
                    ("sjc", "SOJA (SJC)"), ("soy", "SOJA FOB (SOY)")):
        p, c = r.get(f"put_{m}", 0), r.get(f"call_{m}", 0)
        if p or c:
            print(f"{nome}\nPUT: {p}\nCALL: {c}\n")
        else:
            print(f"{nome}\nPUT: 0\nCALL: 0   (nenhuma série listada)\n")
    print(f"Novos registros adicionados: {r['novos']}")
    print(f"Duplicados ignorados: {r['ignorados_duplicidade']}")


def main():
    ap = argparse.ArgumentParser(
        description="Banco histórico da cadeia de opções agro da B3")
    ap.add_argument("--data", help="pregão a coletar (AAAA-MM-DD). "
                                   "Padrão: último dia útil.")
    ap.add_argument("--backfill", nargs=2, metavar=("INICIO", "FIM"),
                    help="preenche um intervalo. Roda só sob comando.")
    ap.add_argument("--dry-run", action="store_true",
                    help="coleta e valida, sem gravar nada")
    ap.add_argument("--dias", type=int, default=5,
                    help="quantos dias úteis recentes verificar em busca de "
                         "lacuna (padrão: 5). A checagem é feita na base "
                         "local; só o que falta é baixado.")
    ap.add_argument("--reprocessar-derivados", action="store_true",
                    help="recalcula as colunas derivadas de todo o histórico "
                         "(mapeamento do futuro, moneyness, liquidez) sem "
                         "baixar nada da B3")
    ap.add_argument("--refazer-radar", action="store_true",
                    help="reconstrói o opcoes_radar.xlsx a partir de TODO o "
                         "histórico em Parquet, sem baixar nada")
    ap.add_argument("--sem-radar", action="store_true",
                    help="não reconstrói o Radar após a coleta")
    ap.add_argument("--refazer-excel", action="store_true",
                    help="reconstrói a janela do Excel a partir da base")
    ap.add_argument("--janela", type=int, default=JANELA_MESES,
                    help=f"meses no Excel (padrão: {JANELA_MESES})")
    ap.add_argument("--sem-excel", action="store_true",
                    help="grava só na base; não mexe no Excel")
    args = ap.parse_args()

    log("=" * 62)
    log(f"Iniciando coleta B3 — opções agro (coletor {ob.VERSAO_COLETOR})")
    if not _motor_parquet():
        log("pyarrow ausente: a base usará CSV comprimido. Para o formato "
            "mais rápido, rode: pip install pyarrow", "AVISO")

    try:
        if args.reprocessar_derivados:
            r = reprocessar_derivados(dry_run=args.dry_run)
            if r and not args.dry_run and not args.sem_excel:
                escrever_excel(args.janela)
            if r and not args.dry_run and not args.sem_radar:
                gerar_radar()
            print(f"\nReprocessamento concluído: {r.get('linhas', 0):,} "
                  f"linha(s), {r.get('mudou', 0):,} célula(s) derivada(s) "
                  f"corrigida(s).")
            print("Nenhuma linha criada ou removida; dados brutos intactos.")
            return 0

        if args.refazer_radar:
            gerar_radar()
            return 0

        if args.refazer_excel:
            escrever_excel(args.janela)
            log("Excel reconstruído a partir da base.")
            return 0

        if args.backfill:
            ini = date.fromisoformat(args.backfill[0])
            fim = date.fromisoformat(args.backfill[1])
            resumos = backfill(ini, fim, dry_run=args.dry_run)
            if not args.dry_run and not args.sem_excel and resumos:
                escrever_excel(args.janela, resumos[-1])
            falhou = [r for r in resumos
                      if r["status"] in ("ERRO_LAYOUT", "REPROVADO_VALIDACAO")]
            return 2 if falhou else 0

        if args.data:
            alvos = [date.fromisoformat(args.data)]
        else:
            alvos = pregoes_faltando(args.dias)
            if not alvos:
                print(f"\nNada a fazer — os {args.dias} último(s) pregão(ões) "
                      f"já constam na base.")
                print("Nenhum download foi feito.")
                return 0
            if len(alvos) > 1:
                log(f"{len(alvos)} pregão(ões) em falta: "
                    f"{', '.join(x.strftime('%d/%m') for x in alvos)}")

        # Do mais antigo para o mais novo: se algo falhar no meio, o que
        # ficou gravado é um trecho contínuo, sem buraco interno.
        resultados = []
        for d in alvos:
            print(f"\nConsultando B3... (pregão de {d:%d/%m/%Y})")
            r = processar(d, dry_run=args.dry_run)
            resultados.append(r)
            if r["status"] == "ERRO_LAYOUT":
                break
            if r["status"] == "OK":
                _relatorio(r)
            if len(alvos) > 1:
                time.sleep(1.5)      # cortesia com o servidor da B3

        r = resultados[-1]
        if len(resultados) > 1:
            tot = sum(x["novos"] for x in resultados)
            log(f"TOTAL DE {len(resultados)} PREGÃO(ÕES): {tot:,} "
                f"registro(s) novo(s)")

        gravados = [x for x in resultados if x["status"] == "OK"]
        novos = sum(x["novos"] for x in gravados)
        ignorados = sum(x["ignorados_duplicidade"] for x in gravados)
        log(f"TOTAL NOVOS REGISTROS: {novos}")
        log(f"DUPLICADOS IGNORADOS: {ignorados}")

        if args.dry_run:
            print("\n[dry-run] Nada foi gravado.")
            return 0

        # O Excel é reescrito se ALGUM pregão entrou, mesmo que outro tenha
        # falhado. Deixar de atualizar a janela por causa de um feriado no
        # meio da lista seria perder o que já foi gravado com sucesso.
        if gravados and not args.sem_excel:
            escrever_excel(args.janela, gravados[-1])
            print(f"\n{EXCEL.name} atualizado com sucesso.")
        # O Radar só é refeito se ALGUM pregão entrou. Coleta que falhou
        # não pode substituir um Radar bom por um gerado de base parcial.
        if gravados and not args.sem_radar:
            gerar_radar()

        ruins = [x for x in resultados
                 if x["status"] in ("ERRO_LAYOUT", "REPROVADO_VALIDACAO")]
        if ruins:
            print(f"\nERRO em {len(ruins)} pregão(ões): "
                  f"{ruins[0]['status']}. Histórico preservado.")
            return 2
        if not gravados:
            print("\nNenhum pregão novo — provavelmente feriado. "
                  "Histórico preservado.")
        return 0

    except KeyboardInterrupt:
        log("interrompido pelo usuário — nada parcial foi gravado", "AVISO")
        return 1
    except Exception as e:
        # Erro é registrado com rastro completo. Engolir exceção aqui
        # seria a forma mais fácil de corromper a base sem ninguém notar.
        log(f"FALHA INESPERADA: {type(e).__name__}: {e}", "ERRO")
        log(traceback.format_exc(), "ERRO")
        print("\nERRO inesperado. Histórico preservado. Veja logs/opcoes_b3.log")
        return 1


if __name__ == "__main__":
    sys.exit(main())
